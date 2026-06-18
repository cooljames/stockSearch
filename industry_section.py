# pyright: reportMissingImports=false
"""
industry_section.py — 네이버 업종별 종목 탐색 섹션 (Standalone Version)
업종 검색(타이핑·초성) → 시가총액순 종목 테이블(종목코드·52주 최저/최고 포함)
→ 선택 종목을 ChartViewer로 시각화, 엑셀(.xlsx) 저장 지원
"""
import csv
import json
import threading
import urllib.request
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
import yfinance as yf
import datetime
import warnings
warnings.filterwarnings("ignore", message=".*possibly delisted; no timezone found.*")
import logging
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

NAVER_INDUSTRY_LIST   = "https://m.stock.naver.com/api/stocks/industry?page=1&pageSize=100"
NAVER_INDUSTRY_STOCKS = (
    "https://m.stock.naver.com/api/stocks/industry/{no}"
    "?page=1&pageSize=100&sortType=marketValue&sortOrder=desc"
)
NAVER_INTEGRATION = "https://m.stock.naver.com/api/stock/{code}/integration"

# 엑셀/CSV 공통 헤더
EXPORT_HEADER = ["순위", "종목명", "종목코드", "업종", "시가총액",
                 "종가", "등락률(%)", "52주최저", "52주최고", "배당금", "배당수익률", "배당기준일", "1년전 배당", "2년전 배당", "3년전 배당", "거래소"]

_CHOSUNG = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"


def _chosung(text: str) -> str:
    """문자열의 한글 음절을 초성으로 변환(영문·기타는 그대로). 초성 검색용."""
    out = []
    for ch in text:
        code = ord(ch)
        if 0xAC00 <= code <= 0xD7A3:
            out.append(_CHOSUNG[(code - 0xAC00) // 588])
        else:
            out.append(ch)
    return "".join(out)


def _fetch_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode("utf-8"))


def _fetch_json_polling(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=10) as r:
        content = r.read()
        try:
            return json.loads(content.decode("cp949"))
        except Exception:
            return json.loads(content.decode("utf-8", errors="ignore"))


def _fetch_realtime_price(code: str) -> tuple:
    """네이버 실시간 폴링 API를 사용하여 실시간 현재가와 등락률을 조회"""
    try:
        url = f"https://polling.finance.naver.com/api/realtime?query=SERVICE_ITEM:{code}"
        res = _fetch_json_polling(url)
        result = res.get("result", {})
        areas = result.get("areas", [])
        if areas and isinstance(areas, list):
            datas = areas[0].get("datas", [])
            if datas and isinstance(datas, list):
                data = datas[0]
                nv = data.get("nv")
                pcv = data.get("pcv")
                cr = data.get("cr")
                if nv is not None and nv > 0:
                    close_price = f"{nv:,}"
                    if pcv is not None and pcv > 0:
                        diff = nv - pcv
                        if diff < 0:
                            ratio = -abs(cr) if cr is not None else 0.0
                        elif diff > 0:
                            ratio = abs(cr) if cr is not None else 0.0
                        else:
                            ratio = 0.0
                    else:
                        ratio = 0.0
                    ratio_str = f"{ratio:+.2f}" if ratio != 0.0 else "0.00"
                    return close_price, ratio_str
    except Exception:
        pass
    return None, None


def _get_annual_dividends(divs) -> dict:
    if divs is None or divs.empty:
        return {}
    
    # 1. Determine frequency
    counts = divs.groupby(divs.index.year).size()
    max_counts = counts.max()
    
    if max_counts >= 8:
        expected_freq = 12
    elif max_counts >= 3:
        expected_freq = 4
    elif max_counts >= 2:
        expected_freq = 2
    else:
        expected_freq = 1
        
    # 2. Impute missing payments for recent years
    annual_sums = {}
    current_year = datetime.datetime.now().year
    
    for year, group in divs.groupby(divs.index.year):
        actual_count = len(group)
        raw_sum = group.sum()
        
        should_impute = False
        if year >= 2023 and year != current_year:
            if expected_freq == 12 and actual_count in (10, 11):
                should_impute = True
            elif expected_freq == 4 and actual_count == 3:
                should_impute = True
                
        if should_impute:
            avg_val = raw_sum / actual_count
            missing_count = expected_freq - actual_count
            annual_sums[year] = raw_sum + (avg_val * missing_count)
        else:
            annual_sums[year] = raw_sum
            
    return annual_sums


def _fetch_naver_details(code: str) -> tuple:
    """Naver API만 사용하여 52주 최저/최고, 배당금, 배당수익률, 배당기준일을 즉시 조회 (매우 빠름)"""
    try:
        d = _fetch_json(NAVER_INTEGRATION.format(code=code))
        
        realtime_price, _ = _fetch_realtime_price(code)
        if realtime_price is not None:
            close_price = realtime_price
        else:
            close_price = d.get("closePrice", "-")
            if close_price == "-" or close_price is None:
                trends = d.get("dealTrendInfos", [])
                if trends and isinstance(trends, list):
                    close_price = trends[0].get("closePrice", "-")
                
        lo = hi = div = div_yield = div_dt = "-"
        for t in d.get("totalInfos", []):
            c = t.get("code")
            if c == "lowPriceOf52Weeks":
                lo = t.get("value", "-")
            elif c == "highPriceOf52Weeks":
                hi = t.get("value", "-")
            elif c == "dividend":
                div = t.get("value", "-")
                div_dt = t.get("valueDesc", "-")
                if div_dt.endswith("."):
                    div_dt = div_dt[:-1]
            elif c == "dividendYieldRatio":
                div_yield = t.get("value", "-")
        
        if lo == "N/A": lo = "-"
        if hi == "N/A": hi = "-"
        if div == "N/A": div = "-"
        if div_yield == "N/A": div_yield = "-"
        if div_dt == "N/A": div_dt = "-"
        
        return lo, hi, div, div_yield, div_dt, close_price
    except Exception:
        return "-", "-", "-", "-", "-", "-"


def _fetch_yfinance_dividends(code: str, close_price: str, div: str, div_yield: str, div_dt: str) -> tuple:
    """yfinance를 사용하여 과거 1~3년 전 배당금을 조회 (상대적으로 느림)"""
    div_y1 = div_y2 = div_y3 = "-"
    base_year = 2025
    if div_dt != "-" and len(div_dt) >= 4 and div_dt[:4].isdigit():
        base_year = int(div_dt[:4])
        
    try:
        ticker = yf.Ticker(f"{code}.KS")
        divs = ticker.dividends
        if divs.empty:
            ticker = yf.Ticker(f"{code}.KQ")
            divs = ticker.dividends
        
        if not divs.empty:
            annual = _get_annual_dividends(divs)
            
            latest_date = divs.index[-1]
            latest_year = latest_date.year
            current_year = datetime.datetime.now().year
            
            if div == "-" and latest_year >= current_year - 1:
                base_year = latest_year
                div_dt = latest_date.strftime("%Y.%m")
                
                val_base = annual.get(base_year, 0)
                div = f"{val_base:,.0f}원" if val_base > 0 else "-"
                
                # Calculate dividend yield from trailing 12-month (TTM) distributions
                try:
                    price_str = str(close_price).replace(",", "")
                    price_val = float(price_str)
                    if price_val > 0:
                        now_tz = datetime.datetime.now(divs.index.tz)
                        one_year_ago = now_tz - datetime.timedelta(days=365)
                        recent_sum = divs[divs.index >= one_year_ago].sum()
                        
                        div_yield_val = (recent_sum / price_val) * 100
                        div_yield = f"{div_yield_val:.2f}%"
                except Exception:
                    pass
            
            y1 = base_year - 1
            y2 = base_year - 2
            y3 = base_year - 3
            
            val_y1 = annual.get(y1, 0)
            val_y2 = annual.get(y2, 0)
            val_y3 = annual.get(y3, 0)
            
            div_y1 = f"{val_y1:,.0f}원" if val_y1 > 0 else "-"
            div_y2 = f"{val_y2:,.0f}원" if val_y2 > 0 else "-"
            div_y3 = f"{val_y3:,.0f}원" if val_y3 > 0 else "-"
    except Exception:
        pass
        
    return div, div_yield, div_dt, div_y1, div_y2, div_y3


def _fetch_details(code: str) -> tuple:
    """단일 종목의 전체 정보(52주 범위 및 1~3년 전 배당금까지) 조회. (동기 실행 및 내보내기용)"""
    lo, hi, div, div_yield, div_dt, close_price = _fetch_naver_details(code)
    div, div_yield, div_dt, div_y1, div_y2, div_y3 = _fetch_yfinance_dividends(code, close_price, div, div_yield, div_dt)
    return lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3



def _fetch_all_sector_stocks(no: int) -> list:
    page = 1
    stocks = []
    while True:
        url = f"https://m.stock.naver.com/api/stocks/industry/{no}?page={page}&pageSize=100&sortType=marketValue&sortOrder=desc"
        data = _fetch_json(url)
        page_stocks = data.get("stocks", [])
        if not page_stocks:
            break
        stocks.extend(page_stocks)
        page += 1
    return stocks


def _fmt_mktcap(raw: int) -> str:
    if raw >= 1_000_000_000_000:
        return f"{raw / 1_000_000_000_000:.1f}조"
    if raw >= 100_000_000:
        return f"{raw / 100_000_000:.0f}억"
    return f"{raw:,}"


def _ex_code(s: dict) -> str:
    ex = s.get("stockExchangeType", {})
    return ex.get("code", "KS") if isinstance(ex, dict) else str(ex)


def _fetch_global_stock_info(code: str, name: str) -> dict:
    """단일 종목의 전체 정보(현재가, 등락률, 시가총액, 거래소, 52주 최저/최고, 배당 정보)를 조회하여 딕셔너리로 반환"""
    try:
        # 1. Naver API
        d = _fetch_json(NAVER_INTEGRATION.format(code=code))
        
        # Parse closePrice & fluctuationsRatio with realtime polling API first
        realtime_price, realtime_ratio = _fetch_realtime_price(code)
        
        if realtime_price is not None:
            close_price = realtime_price
            fluctuations_ratio = realtime_ratio
        else:
            # Parse closePrice
            close_price = d.get("closePrice", "-")
            if close_price == "-" or close_price is None:
                trends = d.get("dealTrendInfos", [])
                if trends and isinstance(trends, list):
                    close_price = trends[0].get("closePrice", "-")
                    
            # Parse fluctuationsRatio
            fluctuations_ratio = d.get("fluctuationsRatio", "0")
            if fluctuations_ratio in ("0", 0, None, "-"):
                trends = d.get("dealTrendInfos", [])
                if trends and isinstance(trends, list):
                    try:
                        curr_price_str = trends[0].get("closePrice", "0").replace(",", "")
                        diff_str = trends[0].get("compareToPreviousClosePrice", "0").replace(",", "")
                        curr_price = float(curr_price_str)
                        diff = float(diff_str)
                        compare_info = trends[0].get("compareToPreviousPrice", {})
                        is_falling = False
                        if isinstance(compare_info, dict):
                            if compare_info.get("name") == "FALLING" or compare_info.get("code") == "5":
                                is_falling = True
                        
                        if is_falling and diff > 0:
                            diff = -diff
                        elif not is_falling and diff < 0:
                            diff = abs(diff)
                            
                        prev_price = curr_price - diff
                        if prev_price > 0:
                            fluctuations_ratio = round((diff / prev_price) * 100, 2)
                        else:
                            fluctuations_ratio = 0.0
                    except Exception:
                        fluctuations_ratio = 0.0
                
        # Parse marketValueRaw
        market_value_raw = d.get("marketValueRaw", 0)
        if not market_value_raw:
            for t in d.get("totalInfos", []):
                if t.get("code") == "marketValue":
                    market_value_raw = t.get("valueRaw", 0)
                    if not market_value_raw:
                        val_str = t.get("value", "")
                        # Parse from Korean units (e.g. "417조 8,848억원")
                        try:
                            val = 0
                            s_clean = val_str.replace(",", "").replace(" ", "").replace("원", "")
                            if "조" in s_clean:
                                parts = s_clean.split("조")
                                val += int(parts[0]) * 1_000_000_000_000
                                if parts[1] and "억" in parts[1]:
                                    val += int(parts[1].replace("억", "")) * 100_000_000
                            elif "억" in s_clean:
                                val += int(s_clean.replace("억", "")) * 100_000_000
                            market_value_raw = val
                        except Exception:
                            market_value_raw = 0
                            
        # Parse stockExchangeType
        exchange = d.get("stockExchangeType", {})
        if not exchange:
            exchange = {"code": "KS"}
            
        # Parse 52 weeks and dividend details
        lo = hi = div = div_yield = div_dt = "-"
        for t in d.get("totalInfos", []):
            c = t.get("code")
            if c == "lowPriceOf52Weeks":
                lo = t.get("value", "-")
            elif c == "highPriceOf52Weeks":
                hi = t.get("value", "-")
            elif c == "dividend":
                div = t.get("value", "-")
                div_dt = t.get("valueDesc", "-")
                if div_dt.endswith("."):
                    div_dt = div_dt[:-1]
            elif c == "dividendYieldRatio":
                div_yield = t.get("value", "-")
                
        # Clean up "N/A"
        if lo == "N/A": lo = "-"
        if hi == "N/A": hi = "-"
        if div == "N/A": div = "-"
        if div_yield == "N/A": div_yield = "-"
        if div_dt == "N/A": div_dt = "-"
        
        # past dividends using yfinance
        div_y1 = div_y2 = div_y3 = "-"
        base_year = 2025
        if div_dt != "-" and len(div_dt) >= 4 and div_dt[:4].isdigit():
            base_year = int(div_dt[:4])
            
        try:
            ticker = yf.Ticker(f"{code}.KS")
            divs = ticker.dividends
            if divs.empty:
                ticker = yf.Ticker(f"{code}.KQ")
                divs = ticker.dividends
            if not divs.empty:
                annual = _get_annual_dividends(divs)
                
                # If Naver dividend is missing (typical for ETFs), fill it using yfinance
                # Only use yfinance data if the latest dividend is recent (within last year or this year)
                latest_date = divs.index[-1]
                latest_year = latest_date.year
                current_year = datetime.datetime.now().year
                
                if div == "-" and latest_year >= current_year - 1:
                    base_year = latest_year
                    div_dt = latest_date.strftime("%Y.%m")
                    
                    val_base = annual.get(base_year, 0)
                    div = f"{val_base:,.0f}원" if val_base > 0 else "-"
                    
                    # Calculate dividend yield from trailing 12-month (TTM) distributions
                    try:
                        price_str = str(close_price).replace(",", "")
                        price_val = float(price_str)
                        if price_val > 0:
                            now_tz = datetime.datetime.now(divs.index.tz)
                            one_year_ago = now_tz - datetime.timedelta(days=365)
                            recent_sum = divs[divs.index >= one_year_ago].sum()
                            
                            div_yield_val = (recent_sum / price_val) * 100
                            div_yield = f"{div_yield_val:.2f}%"
                    except Exception:
                        pass
                
                y1 = base_year - 1
                y2 = base_year - 2
                y3 = base_year - 3
                val_y1 = annual.get(y1, 0)
                val_y2 = annual.get(y2, 0)
                val_y3 = annual.get(y3, 0)
                div_y1 = f"{val_y1:,.0f}원" if val_y1 > 0 else "-"
                div_y2 = f"{val_y2:,.0f}원" if val_y2 > 0 else "-"
                div_y3 = f"{val_y3:,.0f}원" if val_y3 > 0 else "-"
        except Exception:
            pass
            
        return {
            "stockName": name,
            "itemCode": code,
            "closePrice": close_price,
            "fluctuationsRatio": fluctuations_ratio,
            "marketValueRaw": market_value_raw,
            "stockExchangeType": exchange,
            "_low52": lo,
            "_high52": hi,
            "_dividend": div,
            "_div_yield": div_yield,
            "_div_dt": div_dt,
            "_div_y1": div_y1,
            "_div_y2": div_y2,
            "_div_y3": div_y3
        }
    except Exception:
        return {
            "stockName": name,
            "itemCode": code,
            "closePrice": "-",
            "fluctuationsRatio": "0",
            "marketValueRaw": 0,
            "stockExchangeType": {"code": "KS"},
            "_low52": "-",
            "_high52": "-",
            "_dividend": "-",
            "_div_yield": "-",
            "_div_dt": "-",
            "_div_y1": "-",
            "_div_y2": "-",
            "_div_y3": "-"
        }


class IndustrySectorSection:
    """업종 탐색 섹션. build(container) 로 UI 구성."""

    def __init__(self, root: tk.Tk, theme: dict, main_notebook=None, sub_notebook=None):
        self.root           = root
        self.theme          = theme
        self.main_notebook  = main_notebook
        self.sub_notebook   = sub_notebook
        self.bg_card        = theme["bg_card"]
        self.bg_input       = theme["bg_input"]
        self.bg_dark        = theme.get("bg_dark", "#F5F0E8")
        self.text_light     = theme["text_light"]
        self.text_dark      = theme.get("text_dark", "#1A1A1A")
        self.text_muted     = theme["text_muted"]
        self.accent         = theme["accent"]
        self.btn_rss        = theme["btn_rss"]
        self.tertiary_hov   = theme.get("tertiary_hov", "#003DD6")
        self.primary        = theme.get("primary", "#1A1A1A")
        self._groups: list  = []        # 전체 업종 그룹
        self._stocks: list  = []        # 현재 업종 종목
        self._current_group = None
        self._all_display: list = []    # "업종명 (N종목)" 표시 문자열
        self._all_raw: list     = []    # 업종 원본 이름(검색용)
        self._display_to_group: dict = {}
        self._load_token    = 0         # 52주 로딩 stale 방지 토큰
        self._hide_job      = None
        self._sort_col      = None
        self._sort_reverse  = False
        self._global_stocks_cache = {}
        self._global_stocks_exchanges = {}
        self._details_cache = {}

    def _safe_call(self, func, *args, **kwargs):
        """Safely execute a UI update function if the root and main widgets still exist."""
        try:
            if hasattr(self, "root") and self.root and self.root.winfo_exists():
                return func(*args, **kwargs)
        except Exception:
            pass

    def _safe_after(self, delay: int, func, *args, **kwargs):
        """Register a callback with after, but only execute it if the widget/root is not destroyed."""
        def wrapper():
            try:
                if hasattr(self, "root") and self.root and self.root.winfo_exists():
                    func(*args, **kwargs)
            except Exception:
                pass
        try:
            if hasattr(self, "root") and self.root and self.root.winfo_exists():
                self.root.after(delay, wrapper)
        except Exception:
            pass

    def build(self, container: tk.Frame) -> None:
        frame = tk.LabelFrame(
            container,
            text="  📊 업종별 종목 탐색 (한국)  ",
            font=("맑은 고딕", 10, "bold"),
            bg=self.bg_card, fg=self.text_dark,
            bd=3, relief="solid", padx=10, pady=10,
        )
        frame.pack(fill="both", expand=True, pady=4)
        self._frame = frame

        # ── 1행: 업종 검색창 + 새로고침 + 통계 ──
        top_row = tk.Frame(frame, bg=self.bg_card)
        top_row.pack(fill="x", pady=(0, 6))

        tk.Label(top_row, text="업종:",
                 font=("맑은 고딕", 9, "bold"),
                 bg=self.bg_card, fg=self.text_dark).pack(side="left", padx=(0, 4))

        # 타이핑·초성 검색 가능한 입력창
        self.industry_var = tk.StringVar()
        self.industry_entry = ttk.Entry(
            top_row, textvariable=self.industry_var,
            width=24, font=("맑은 고딕", 9),
        )
        self.industry_entry.pack(side="left", padx=(0, 8))
        self.industry_entry.bind("<KeyRelease>", self._on_industry_type)
        self.industry_entry.bind("<Down>",   self._focus_listbox)
        self.industry_entry.bind("<Return>", lambda e: self._commit_first())
        self.industry_entry.bind("<Escape>", lambda e: self._hide_popup())
        self.industry_entry.bind("<FocusIn>",  lambda e: self._on_industry_type(e))
        self.industry_entry.bind("<FocusOut>", self._schedule_hide)

        # 업종 선택 버튼
        self.select_industry_btn = tk.Button(
            top_row, text="📂 업종 선택",
            font=("맑은 고딕", 9, "bold"), bg=self.accent, fg=self.text_dark,
            activebackground=self.theme.get("hover_dark", "#E6B800"),
            bd=1, relief="solid", cursor="hand2", padx=8, pady=3,
            command=self._open_industry_selector,
        )
        self.select_industry_btn.pack(side="left", padx=(0, 8))

        self.refresh_btn = tk.Button(
            top_row, text="🔄 전체 종목",
            font=("맑은 고딕", 9), bg=self.primary, fg=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=8, pady=4,
            command=self._load_all_stocks_thread,
        )
        self.refresh_btn.pack(side="left", padx=(0, 12))

        self.stats_lbl = tk.Label(
            top_row, text="업종을 불러오는 중...",
            font=("맑은 고딕", 8), bg=self.bg_card, fg=self.text_muted,
        )
        self.stats_lbl.pack(side="left")
        # 주식 검색 입력 및 버튼
        self.stock_search_var = tk.StringVar()
        self.stock_search_entry = ttk.Entry(
            top_row, textvariable=self.stock_search_var,
            width=14, font=("맑은 고딕", 9),
        )
        self.stock_search_entry.pack(side="left", padx=(12, 4))
        self.stock_search_entry.bind("<Return>", lambda e: self._search_stocks())
        self.stock_search_entry.bind("<Control-Return>", lambda e: self._global_search_stocks_thread())
        self.search_btn = tk.Button(
            top_row, text="🔍 검색",
            font=("맑은 고딕", 9), bg=self.primary, fg=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=6, pady=4,
            command=self._search_stocks,
        )
        self.search_btn.pack(side="left")
        self.global_search_btn = tk.Button(
            top_row, text="🔍 전체 검색 (Ctrl+Enter)",
            font=("맑은 고딕", 9, "bold"), bg=self.accent, fg=self.text_dark,
            activebackground=self.theme.get("hover_dark", "#E6B800"),
            bd=1, relief="solid", cursor="hand2", padx=8, pady=3,
            command=self._global_search_stocks_thread,
        )
        self.global_search_btn.pack(side="left", padx=(4, 0))
        # 검색 자동완성 팝업(Listbox) — frame 위에 place 로 떠 있음
        self._popup = tk.Frame(frame, bg=self.text_dark, bd=1, relief="solid")
        self.listbox = tk.Listbox(
            self._popup, height=10, width=30, font=("맑은 고딕", 9),
            activestyle="none", exportselection=False,
            bg=self.bg_input, fg=self.text_dark,
            selectbackground=self.accent, selectforeground=self.text_light,
            highlightthickness=0, bd=0,
        )
        self.listbox.pack(fill="both", expand=True)
        self.listbox.bind("<ButtonRelease-1>", lambda e: self._commit_listbox())
        self.listbox.bind("<Return>",          lambda e: self._commit_listbox())
        self.listbox.bind("<Escape>",          lambda e: self._hide_popup())

        # ── 2행: 종목 Treeview ──
        tree_frame = tk.Frame(frame, bg=self.bg_card)
        tree_frame.pack(fill="both", expand=True, pady=(0, 6))

        col_cfg = [
            ("순위",     40, "center"),
            ("종목명",   130, "w"),
            ("종목코드",  70, "center"),
            ("시가총액",  85, "e"),
            ("종가",      80, "e"),
            ("등락률",    72, "e"),
            ("52주최저",  80, "e"),
            ("52주최고",  80, "e"),
            ("배당금",    80, "e"),
            ("배당수익률", 80, "e"),
            ("배당기준일", 85, "center"),
            ("1년전 배당", 80, "e"),
            ("2년전 배당", 80, "e"),
            ("3년전 배당", 80, "e"),
            ("거래소",    52, "center"),
        ]
        cols = tuple(c[0] for c in col_cfg)
        self.tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings",
            selectmode="extended", height=8,
        )
        for col, width, anchor in col_cfg:
            if col == "배당수익률":
                self.tree.heading(col, text=col, command=self._toggle_dividend_yield_sort)
            else:
                self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor=anchor, stretch=False)

        self.vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self._on_tree_scroll)
        self.tree.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")

        self.tree.tag_configure("rise", foreground="#E63B2E")
        self.tree.tag_configure("fall", foreground="#0055FF")
        self.tree.tag_configure("odd",  background="#F7F3EC")

        # ── 3행: 선택·실행 버튼 + 엑셀 저장 ──
        btn_row = tk.Frame(frame, bg=self.bg_card)
        btn_row.pack(fill="x")

        for text, cmd, padx in [
            ("상위 5 선택", self._select_top5,  (0, 6)),
            ("전체 선택",   self._select_all,    (0, 6)),
            ("선택 해제",   self._deselect_all,  (0, 16)),
        ]:
            tk.Button(
                btn_row, text=text,
                font=("맑은 고딕", 9), bg=self.bg_input, fg=self.text_dark,
                bd=1, relief="solid", cursor="hand2", padx=8, pady=5,
                command=cmd,
            ).pack(side="left", padx=padx)

        self.analyze_btn = tk.Button(
            btn_row,
            text="📈  선택 종목 차트 조회",
            font=("맑은 고딕", 10, "bold"),
            bg=self.btn_rss, fg=self.text_light,
            activebackground=self.tertiary_hov, activeforeground=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=16, pady=6,
            command=self._show_chart,
        )
        self.analyze_btn.pack(side="left")

        # 오른쪽 맨끝: 엑셀 저장(현재 업종 / 전체 업종)
        self.export_mb = tk.Menubutton(
            btn_row, text="💾 엑셀로 저장 ▾",
            font=("맑은 고딕", 9, "bold"),
            bg=self.primary, fg=self.text_light,
            activebackground=self.tertiary_hov, activeforeground=self.text_light,
            bd=0, relief="flat", cursor="hand2", padx=12, pady=6,
        )
        export_menu = tk.Menu(self.export_mb, tearoff=0, font=("맑은 고딕", 9))
        export_menu.add_command(label="현재 업종 종목 저장",
                                command=lambda: self._export("current"))
        export_menu.add_command(label="전체 업종 종목 저장",
                                command=lambda: self._export("all"))
        self.export_mb["menu"] = export_menu
        self.export_mb.pack(side="right")

        # 초기 업종 목록 로드
        self._load_industries_thread()

    # ──────────────────────────────────────────────
    # 업종 목록 로드
    # ──────────────────────────────────────────────
    def _load_industries_thread(self):
        self.refresh_btn.configure(state="disabled", text="로딩 중...")
        threading.Thread(target=self._load_industries, daemon=True).start()

    def _load_industries(self):
        try:
            data   = _fetch_json(NAVER_INDUSTRY_LIST)
            groups = data.get("groups", [])
            self._groups = sorted(groups, key=lambda g: g["name"])
            self.root.after(0, self._init_industry_list)
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror(
                "업종 목록 오류", f"업종 목록 조회 실패:\n{msg}"))
        finally:
            self.root.after(0, lambda: self.refresh_btn.configure(
                state="normal", text="🔄 전체 종목"))

    def _load_all_stocks_thread(self):
        self.refresh_btn.configure(state="disabled", text="로딩 중...")
        self.stats_lbl.configure(text="전체 종목 불러오는 중...", fg=self.text_muted)
        threading.Thread(target=self._load_all_stocks, daemon=True).start()

    def _load_all_stocks(self):
        try:
            # 캐시가 빈 경우 구축
            if not self._global_stocks_cache:
                self._build_global_stocks_cache()
            
            matched_stocks = list(self._global_stocks_cache.items())
            if not matched_stocks:
                self.root.after(0, lambda: messagebox.showwarning(
                    "대기 필요", "종목 정보를 불러오는 중입니다. 잠시 후 다시 시도해 주세요."))
                return

            self._set_stats(f"실시간 시세 가져오는 중 (0/{len(matched_stocks)})...", self.text_muted)
            
            # 실시간 시세 일괄 조회 (100종목 단위 청크)
            codes = [code for code, name in matched_stocks]
            realtime_data = {}
            
            chunk_size = 100
            chunks = [codes[i:i + chunk_size] for i in range(0, len(codes), chunk_size)]
            
            completed_count = 0
            
            def fetch_chunk(chunk_codes):
                nonlocal completed_count
                try:
                    url = "https://polling.finance.naver.com/api/realtime?query=SERVICE_ITEM:" + ",".join(chunk_codes)
                    res = _fetch_json_polling(url)
                    result = res.get("result", {})
                    areas = result.get("areas", [])
                    chunk_res = {}
                    if areas and isinstance(areas, list):
                        datas = areas[0].get("datas", [])
                        for data in datas:
                            cd = data.get("cd")
                            nv = data.get("nv")
                            pcv = data.get("pcv")
                            cr = data.get("cr")
                            count_listed = data.get("countOfListedStock", 0)
                            
                            mkt_val = 0
                            if nv is not None and count_listed:
                                try:
                                    mkt_val = int(nv) * int(count_listed)
                                except Exception:
                                    pass
                                    
                            if cd and nv is not None and nv > 0:
                                close_price = f"{nv:,}"
                                if pcv is not None and pcv > 0:
                                    diff = nv - pcv
                                    if diff < 0:
                                        ratio = -abs(cr) if cr is not None else 0.0
                                    elif diff > 0:
                                        ratio = abs(cr) if cr is not None else 0.0
                                    else:
                                        ratio = 0.0
                                else:
                                    ratio = 0.0
                                ratio_str = f"{ratio:+.2f}" if ratio != 0.0 else "0.00"
                                chunk_res[cd] = (close_price, ratio_str, mkt_val)
                    return chunk_res
                except Exception:
                    return {}

            final_stocks = []
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = [executor.submit(fetch_chunk, chunk) for chunk in chunks]
                for fut in as_completed(futures):
                    res = fut.result()
                    realtime_data.update(res)
                    completed_count += len(res)
                    self._set_stats(f"실시간 시세 가져오는 중 ({completed_count}/{len(matched_stocks)})...", self.text_muted)
            
            exchanges_cache = getattr(self, "_global_stocks_exchanges", {})
            
            for code, name in matched_stocks:
                close_price, fluctuations_ratio, mkt_val = realtime_data.get(code, ("-", "0.00", 0))
                ex_code_val = exchanges_cache.get(code, "KS")
                final_stocks.append({
                    "stockName": name,
                    "itemCode": code,
                    "closePrice": close_price,
                    "fluctuationsRatio": fluctuations_ratio,
                    "marketValueRaw": mkt_val,
                    "stockExchangeType": {"code": ex_code_val},
                })
            
            final_stocks = sorted(final_stocks, key=lambda s: s.get("stockName", ""))
            
            self._stocks = final_stocks
            self._all_stocks = list(final_stocks)
            self._current_group = None
            self._sort_col = None
            self._sort_reverse = False
            
            self.root.after(0, lambda: self._populate_tree(group=None))
            self.root.after(0, lambda: self.stats_lbl.configure(text=f"📊 전체 종목: 총 {len(final_stocks)}종목", fg=self.text_dark))
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("전체 종목 오류", f"전체 종목을 불러오는 중 오류 발생:\n{msg}"))
            self.root.after(0, lambda msg=err_msg: self.stats_lbl.configure(text=f"오류: {msg}", fg="#E63B2E"))
        finally:
            self.root.after(0, lambda: self.refresh_btn.configure(state="normal", text="🔄 전체 종목"))

    def _init_industry_list(self):
        self._all_raw     = [g["name"] for g in self._groups]
        self._all_display = [f"{g['name']}  ({g['totalCount']}종목)"
                             for g in self._groups]
        self._display_to_group = dict(zip(self._all_display, self._groups))
        if self._groups:
            self.industry_var.set(self._all_display[0])
            self._load_for_group(self._groups[0])
            # 백그라운드에서 전체 종목(이름, 코드) 로컬 캐시 구축 시작
            threading.Thread(target=self._build_global_stocks_cache, daemon=True).start()

    def _build_global_stocks_cache(self):
        if not self._groups:
            return
        temp_cache = {}
        temp_exchanges = {}
        def fetch_group_stocks(g):
            try:
                return _fetch_all_sector_stocks(g["no"])
            except Exception:
                return []
        try:
            with ThreadPoolExecutor(max_workers=16) as executor:
                # Map futures to their groups
                futures = {executor.submit(fetch_group_stocks, g): g for g in self._groups}
                
                # Fetch all ETFs in parallel (pages 1 to 15)
                etf_futures = []
                for p in range(1, 16):
                    etf_url = f"https://m.stock.naver.com/api/stocks/etf?page={p}&pageSize=100"
                    etf_futures.append(executor.submit(_fetch_json, etf_url))
                
                for fut in as_completed(futures):
                    g = futures[fut]
                    stocks = fut.result()
                    
                    # Calculate stats manually
                    def get_ratio(s):
                        val = s.get("fluctuationsRatio", 0)
                        try:
                            return float(val)
                        except (ValueError, TypeError):
                            return 0.0
                            
                    g['totalCount'] = len(stocks)
                    g['riseCount'] = sum(1 for s in stocks if get_ratio(s) > 0)
                    g['fallCount'] = sum(1 for s in stocks if get_ratio(s) < 0)
                    g['steadyCount'] = sum(1 for s in stocks if get_ratio(s) == 0)
                    ratios = [get_ratio(s) for s in stocks]
                    g['changeRate'] = sum(ratios) / len(ratios) if ratios else 0.0
                    
                    for s in stocks:
                        code = s.get("itemCode")
                        name = s.get("stockName")
                        if code and name and code.isdigit() and len(code) == 6:
                            temp_cache[code] = name
                            temp_exchanges[code] = _ex_code(s)
                            
                for fut in as_completed(etf_futures):
                    try:
                        res = fut.result()
                        etf_stocks = res.get("stocks", [])
                        for s in etf_stocks:
                            code = s.get("itemCode")
                            name = s.get("stockName")
                            if code and name and code.isdigit() and len(code) == 6:
                                temp_cache[code] = name
                                temp_exchanges[code] = _ex_code(s)
                    except Exception:
                        pass
                        
            self._global_stocks_cache = temp_cache
            self._global_stocks_exchanges = temp_exchanges
            # Rebuild display names and trigger UI update
            self.root.after(0, self._update_industry_list_ui)
        except Exception:
            pass

    def _update_industry_list_ui(self):
        selected_group_name = None
        if self._current_group:
            selected_group_name = self._current_group["name"]
            
        self._all_display = [f"{g['name']}  ({g['totalCount']}종목)"
                             for g in self._groups]
        self._display_to_group = dict(zip(self._all_display, self._groups))
        
        # If the currently selected group is in the new list, update the entry display
        if selected_group_name:
            for display in self._all_display:
                group = self._display_to_group.get(display)
                if group and group["name"] == selected_group_name:
                    self.industry_var.set(display)
                    # Also update the stats label if it was showing stats for the current group
                    self._populate_stats(group)
                    break

    def _open_industry_selector(self) -> None:
        if not self._all_display:
            messagebox.showwarning("업종 선택", "업종 목록이 아직 로드되지 않았거나 비어 있습니다. 잠시 후 다시 시도해 주세요.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("업종 선택")
        dialog.geometry("380x500")
        dialog.configure(bg=self.bg_dark)
        dialog.resizable(False, False)
        
        # Center dialog relative to root
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Position relative to root window
        rx = self.root.winfo_rootx()
        ry = self.root.winfo_rooty()
        rw = self.root.winfo_width()
        rh = self.root.winfo_height()
        x = rx + (rw - 380) // 2
        y = ry + (rh - 500) // 2
        dialog.geometry(f"+{x}+{y}")

        # Border frame to match Neo-Brutalist design
        main_frame = tk.Frame(dialog, bg=self.bg_dark, bd=2, relief="solid")
        main_frame.pack(fill="both", expand=True, padx=12, pady=12)

        # Header Title
        title_lbl = tk.Label(
            main_frame, text="📁 업종 선택",
            font=("맑은 고딕", 12, "bold"),
            bg=self.accent, fg=self.text_dark,
            bd=2, relief="solid", pady=6
        )
        title_lbl.pack(fill="x", padx=10, pady=(10, 8))

        # Search box
        search_frame = tk.Frame(main_frame, bg=self.bg_dark)
        search_frame.pack(fill="x", padx=10, pady=4)
        
        tk.Label(
            search_frame, text="검색어:",
            font=("맑은 고딕", 9, "bold"),
            bg=self.bg_dark, fg=self.text_dark
        ).pack(side="left", padx=(0, 6))

        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, font=("맑은 고딕", 9))
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.focus_set()

        # Listbox for industries
        list_frame = tk.Frame(main_frame, bg=self.bg_dark, bd=2, relief="solid")
        list_frame.pack(fill="both", expand=True, padx=10, pady=8)

        dialog_listbox = tk.Listbox(
            list_frame, font=("맑은 고딕", 9),
            activestyle="none", exportselection=False,
            bg=self.bg_input, fg=self.text_dark,
            selectbackground=self.accent, selectforeground=self.text_dark,
            highlightthickness=0, bd=0,
        )
        dialog_listbox.pack(side="left", fill="both", expand=True)

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=dialog_listbox.yview)
        dialog_listbox.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")

        # Function to filter the list
        def update_list(*args):
            query = search_var.get().strip()
            dialog_listbox.delete(0, "end")
            matches = self._matches(query)
            for m in matches:
                dialog_listbox.insert("end", m)
            if dialog_listbox.size() > 0:
                dialog_listbox.selection_set(0)

        # Initial population
        update_list()

        search_var.trace_add("write", update_list)

        # Confirm and close functions
        def confirm():
            sel = dialog_listbox.curselection()
            if sel:
                val = dialog_listbox.get(sel[0])
                self._commit_display(val)
                dialog.destroy()

        def cancel():
            dialog.destroy()

        # Button Row
        btn_frame = tk.Frame(main_frame, bg=self.bg_dark)
        btn_frame.pack(fill="x", padx=10, pady=(4, 10))

        cancel_btn = tk.Button(
            btn_frame, text="취소",
            font=("맑은 고딕", 9), bg=self.bg_input, fg=self.text_dark,
            bd=1, relief="solid", cursor="hand2", padx=12, pady=4,
            command=cancel
        )
        cancel_btn.pack(side="right", padx=(6, 0))

        confirm_btn = tk.Button(
            btn_frame, text="선택 완료",
            font=("맑은 고딕", 9, "bold"), bg=self.accent, fg=self.text_dark,
            bd=2, relief="solid", cursor="hand2", padx=16, pady=4,
            command=confirm
        )
        confirm_btn.pack(side="right")

        # Bind keys
        dialog_listbox.bind("<Double-Button-1>", lambda e: confirm())
        search_entry.bind("<Return>", lambda e: confirm())
        dialog_listbox.bind("<Return>", lambda e: confirm())
        dialog.bind("<Escape>", lambda e: cancel())


    # ──────────────────────────────────────────────
    # 업종 검색(타이핑·초성) 자동완성
    # ──────────────────────────────────────────────
    def _matches(self, typed: str) -> list:
        typed = typed.strip()
        if not typed:
            return list(self._all_display)
        q  = typed.lower()
        has_jamo     = any(0x3131 <= ord(c) <= 0x314E for c in typed)
        has_syllable = any(0xAC00 <= ord(c) <= 0xD7A3 for c in typed)
        chosung_mode = has_jamo and not has_syllable
        qc  = _chosung(typed)
        out = []
        for disp, raw in zip(self._all_display, self._all_raw):
            if q in raw.lower():
                out.append(disp)
            elif chosung_mode and qc in _chosung(raw):
                out.append(disp)
        return out

    def _on_industry_type(self, event):
        keysym = getattr(event, "keysym", "")
        if keysym in ("Up", "Down", "Return", "Escape", "Tab",
                       "Left", "Right", "Shift_L", "Shift_R"):
            return
        if not self._all_display:
            return
        matches = self._matches(self.industry_var.get())
        self._show_popup(matches)

    def _show_popup(self, matches: list):
        if self._hide_job:
            self.root.after_cancel(self._hide_job)
            self._hide_job = None
        if not matches:
            self._hide_popup()
            return
        self.listbox.delete(0, "end")
        for m in matches:
            self.listbox.insert("end", m)
        self.listbox.configure(height=min(10, len(matches)))
        self._popup.place(in_=self.industry_entry, x=0, rely=1.0, y=2, anchor="nw")
        self._popup.lift()

    def _hide_popup(self):
        self._popup.place_forget()

    def _schedule_hide(self, _event):
        self._hide_job = self.root.after(180, self._hide_popup)

    def _focus_listbox(self, _event):
        if self._popup.winfo_ismapped() and self.listbox.size():
            self.listbox.focus_set()
            self.listbox.selection_clear(0, "end")
            self.listbox.selection_set(0)
            self.listbox.activate(0)
        return "break"

    def _commit_first(self):
        if self.listbox.size():
            self._commit_display(self.listbox.get(0))

    def _commit_listbox(self):
        sel = self.listbox.curselection()
        if sel:
            self._commit_display(self.listbox.get(sel[0]))

    def _commit_display(self, display: str):
        self.industry_var.set(display)
        self._hide_popup()
        self.industry_entry.icursor("end")
        group = self._display_to_group.get(display)
        if group:
            self.industry_entry.focus_set()
            self._load_for_group(group)

    # ──────────────────────────────────────────────
    # 업종 선택 시 종목 로드
    # ──────────────────────────────────────────────
    def _load_for_group(self, group: dict):
        self._current_group = group
        self.stats_lbl.configure(
            text=f"{group['name']} 종목 로딩 중...", fg=self.text_muted)
        threading.Thread(
            target=self._load_sector_stocks, args=(group,), daemon=True
        ).start()

    def _load_sector_stocks(self, group: dict):
        try:
            stocks = _fetch_all_sector_stocks(group["no"])
            stocks = sorted(stocks,
                            key=lambda s: int(s.get("marketValueRaw", 0)),
                            reverse=True)
            self._stocks = stocks
            self._all_stocks = list(stocks)
            self._sort_col = None
            self._sort_reverse = False
            
            # Calculate stats manually
            def get_ratio(s):
                val = s.get("fluctuationsRatio", 0)
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 0.0
                    
            group['totalCount'] = len(stocks)
            group['riseCount'] = sum(1 for s in stocks if get_ratio(s) > 0)
            group['fallCount'] = sum(1 for s in stocks if get_ratio(s) < 0)
            group['steadyCount'] = sum(1 for s in stocks if get_ratio(s) == 0)
            ratios = [get_ratio(s) for s in stocks]
            group['changeRate'] = sum(ratios) / len(ratios) if ratios else 0.0
            
            self._safe_after(0, lambda g=group: self._populate_tree(g))
            self._safe_after(0, lambda g=group: self._update_single_group_in_displays(g))
        except Exception as e:
            err_msg = str(e)
            self._safe_after(0, lambda msg=err_msg: self.stats_lbl.configure(
                text=f"오류: {msg}", fg="#E63B2E") if hasattr(self, "stats_lbl") and self.stats_lbl and self.stats_lbl.winfo_exists() else None)

    def _update_single_group_in_displays(self, group: dict):
        try:
            if not hasattr(self, "root") or not self.root or not self.root.winfo_exists():
                return
        except Exception:
            return
        self._all_display = [f"{g['name']}  ({g['totalCount']}종목)"
                             for g in self._groups]
        self._display_to_group = dict(zip(self._all_display, self._groups))
        for display in self._all_display:
            g = self._display_to_group.get(display)
            if g and g["no"] == group["no"]:
                try:
                    if hasattr(self, "industry_var") and self.industry_var:
                        self.industry_var.set(display)
                except Exception:
                    pass
                break

    def _populate_tree(self, group: dict = None):
        try:
            if not hasattr(self, "tree") or not self.tree or not self.tree.winfo_exists():
                return
        except Exception:
            return
        self._update_headers()
        for row in self.tree.get_children():
            self.tree.delete(row)

        if group:
            cr    = float(group.get("changeRate", 0))
            sign  = "▲" if cr >= 0 else "▼"
            color = "#E63B2E" if cr >= 0 else "#0055FF"
            self.stats_lbl.configure(
                fg=color,
                text=(
                    f"{sign} {abs(cr):.2f}%  |  "
                    f"상승 {group['riseCount']}  하락 {group['fallCount']}  "
                    f"보합 {group['steadyCount']}  (총 {group['totalCount']}종목)"
                ),
            )
        else:
            self.stats_lbl.configure(
                fg=self.text_dark,
                text=f"🔍 전체 검색 결과: 총 {len(self._stocks)}종목",
            )

        for i, s in enumerate(self._stocks):
            mkt_raw = int(s.get("marketValueRaw", 0))
            ratio   = float(s.get("fluctuationsRatio", 0))

            tags = []
            if ratio > 0:   tags.append("rise")
            elif ratio < 0: tags.append("fall")
            if i % 2 == 1:  tags.append("odd")

            if ratio > 0:
                arrow = "▲"
                ratio_str = f"{arrow} {abs(ratio):.2f}%"
            elif ratio < 0:
                arrow = "▼"
                ratio_str = f"{arrow} -{abs(ratio):.2f}%"
            else:
                ratio_str = "0.00%"

            self.tree.insert(
                "", "end", iid=str(i),
                values=(
                    i + 1,
                    s.get("stockName", ""),
                    s.get("itemCode", ""),
                    _fmt_mktcap(mkt_raw),
                    s.get("closePrice", ""),
                    ratio_str,
                    s.get("_low52", "…"),
                    s.get("_high52", "…"),
                    s.get("_dividend", "…"),
                    s.get("_div_yield", "…"),
                    s.get("_div_dt", "…"),
                    s.get("_div_y1", "…"),
                    s.get("_div_y2", "…"),
                    s.get("_div_y3", "…"),
                    _ex_code(s),
                ),
                tags=tuple(tags),
            )

        self._load_token += 1
        token = self._load_token
        threading.Thread(target=self._fetch_details_all,
                         args=(self._stocks, token), daemon=True).start()

        if len(self._stocks) > 150:
            self.root.after(100, self._lazy_load_current_viewport)

    def _fetch_details_all(self, stocks: list, token: int):
        # 1. Check cache first
        for i, s in enumerate(stocks):
            code = s.get("itemCode", "")
            if code in self._details_cache:
                c = self._details_cache[code]
                s["_low52"] = c["lo"]
                s["_high52"] = c["hi"]
                s["_dividend"] = c["div"]
                s["_div_yield"] = c["div_yield"]
                s["_div_dt"] = c["div_dt"]
                s["_div_y1"] = c["div_y1"]
                s["_div_y2"] = c["div_y2"]
                s["_div_y3"] = c["div_y3"]
                self.root.after(0, self._update_details_row, i, c["lo"], c["hi"], c["div"], c["div_yield"], c["div_dt"], c["div_y1"], c["div_y2"], c["div_y3"], token)

        # 2. Identify targets for Naver API (Phase 1)
        targets_naver = [(i, s) for i, s in enumerate(stocks) if s.get("itemCode", "") not in self._details_cache]
        if not targets_naver:
            return

        # Optimization: Skip background details fetching if there are too many stocks (avoid rate limits/freezing)
        if len(stocks) > 150:
            return

        def work_naver(i_s):
            i, s = i_s
            code = s.get("itemCode", "")
            lo, hi, div, div_yield, div_dt, close_price = _fetch_naver_details(code)
            return i, code, lo, hi, div, div_yield, div_dt, close_price

        # Phase 1: Fetch Naver details (lo, hi, div, div_yield, div_dt) in parallel
        naver_results = []
        try:
            with ThreadPoolExecutor(max_workers=16) as ex:
                futs = [ex.submit(work_naver, t) for t in targets_naver]
                for fut in as_completed(futs):
                    if token != self._load_token:
                        return
                    i, code, lo, hi, div, div_yield, div_dt, close_price = fut.result()
                    
                    stocks[i]["_low52"]  = lo
                    stocks[i]["_high52"] = hi
                    stocks[i]["_dividend"] = div
                    stocks[i]["_div_yield"] = div_yield
                    stocks[i]["_div_dt"] = div_dt
                    
                    # Update row immediately with Naver details (and "…" for yfinance)
                    self.root.after(0, self._update_details_row, i, lo, hi, div, div_yield, div_dt, "…", "…", "…", token)
                    naver_results.append((i, code, lo, hi, div, div_yield, div_dt, close_price))
        except Exception:
            pass

        # Phase 2: Fetch yfinance details (div_y1, div_y2, div_y3) for those stocks
        def work_yfinance(res):
            i, code, lo, hi, div, div_yield, div_dt, close_price = res
            div, div_yield, div_dt, div_y1, div_y2, div_y3 = _fetch_yfinance_dividends(code, close_price, div, div_yield, div_dt)
            return i, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3

        try:
            # Using 4 workers for yfinance to be gentle
            with ThreadPoolExecutor(max_workers=4) as ex:
                futs = [ex.submit(work_yfinance, r) for r in naver_results]
                for fut in as_completed(futs):
                    if token != self._load_token:
                        return
                    i, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3 = fut.result()
                    
                    stocks[i]["_low52"]  = lo
                    stocks[i]["_high52"] = hi
                    stocks[i]["_dividend"] = div
                    stocks[i]["_div_yield"] = div_yield
                    stocks[i]["_div_dt"] = div_dt
                    stocks[i]["_div_y1"] = div_y1
                    stocks[i]["_div_y2"] = div_y2
                    stocks[i]["_div_y3"] = div_y3
                    
                    # Cache the result
                    self._details_cache[code] = {
                        "lo": lo, "hi": hi,
                        "div": div, "div_yield": div_yield, "div_dt": div_dt,
                        "div_y1": div_y1, "div_y2": div_y2, "div_y3": div_y3
                    }
                    
                    # Update row with final values
                    self.root.after(0, self._update_details_row, i, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3, token)
        except Exception:
            pass

    def _update_details_row(self, i, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3, token):
        if token != self._load_token:
            return
        try:
            if not hasattr(self, "tree") or not self.tree or not self.tree.winfo_exists():
                return
            iid = str(i)
            if self.tree.exists(iid):
                self.tree.set(iid, "52주최저", lo)
                self.tree.set(iid, "52주최고", hi)
                self.tree.set(iid, "배당금", div)
                self.tree.set(iid, "배당수익률", div_yield)
                self.tree.set(iid, "배당기준일", div_dt)
                self.tree.set(iid, "1년전 배당", div_y1)
                self.tree.set(iid, "2년전 배당", div_y2)
                self.tree.set(iid, "3년전 배당", div_y3)
        except Exception:
            pass

    def _on_tree_scroll(self, first, last):
        if hasattr(self, "vsb"):
            try:
                if self.vsb.winfo_exists():
                    self.vsb.set(first, last)
            except Exception:
                pass
            
        # Debounce the viewport calculations and data fetching to avoid high-frequency API calls
        if hasattr(self, "_lazy_load_job") and self._lazy_load_job:
            try:
                if self.root.winfo_exists():
                    self.root.after_cancel(self._lazy_load_job)
            except Exception:
                pass
            self._lazy_load_job = None
            
        self._lazy_load_job = self._safe_after(150, lambda: self._lazy_load_visible(float(first), float(last)))

    def _lazy_load_current_viewport(self):
        try:
            first, last = self.tree.yview()
            self._lazy_load_visible(first, last)
        except Exception:
            pass

    def _lazy_load_visible(self, first, last):
        self._lazy_load_job = None
        total = len(self._stocks)
        if total == 0:
            return
            
        # Calculate indexes visible, add a small padding of 5
        start_idx = max(0, int(first * total) - 5)
        end_idx = min(total - 1, int(last * total) + 5)
        
        targets = []
        token = self._load_token
        
        # Update from cache first, collect non-cached
        for idx in range(start_idx, end_idx + 1):
            s = self._stocks[idx]
            code = s.get("itemCode", "")
            if not code:
                continue
            if code in self._details_cache:
                c = self._details_cache[code]
                s["_low52"] = c["lo"]
                s["_high52"] = c["hi"]
                s["_dividend"] = c["div"]
                s["_div_yield"] = c["div_yield"]
                s["_div_dt"] = c["div_dt"]
                s["_div_y1"] = c["div_y1"]
                s["_div_y2"] = c["div_y2"]
                s["_div_y3"] = c["div_y3"]
                self._update_details_row(idx, c["lo"], c["hi"], c["div"], c["div_yield"], c["div_dt"], c["div_y1"], c["div_y2"], c["div_y3"], token)
            else:
                targets.append((idx, s))
                
        if not targets:
            return
            
        # Fetch non-cached ones in background
        threading.Thread(target=self._fetch_batch_details, args=(targets, token), daemon=True).start()

    def _fetch_batch_details(self, targets: list, token: int):
        if token != self._load_token:
            return

        def work_naver(idx_s):
            idx, s = idx_s
            code = s.get("itemCode", "")
            lo, hi, div, div_yield, div_dt, close_price = _fetch_naver_details(code)
            return idx, code, lo, hi, div, div_yield, div_dt, close_price

        # Phase 1: Fetch Naver details in parallel
        naver_results = []
        try:
            with ThreadPoolExecutor(max_workers=8) as ex:
                futs = [ex.submit(work_naver, t) for t in targets]
                for fut in as_completed(futs):
                    if token != self._load_token:
                        return
                    try:
                        idx, code, lo, hi, div, div_yield, div_dt, close_price = fut.result()
                        # Update the stock dict immediately
                        self._stocks[idx]["_low52"] = lo
                        self._stocks[idx]["_high52"] = hi
                        self._stocks[idx]["_dividend"] = div
                        self._stocks[idx]["_div_yield"] = div_yield
                        self._stocks[idx]["_div_dt"] = div_dt
                        
                        # Update row immediately with Naver details (and "…" for yfinance)
                        self.root.after(0, self._update_details_row, idx, lo, hi, div, div_yield, div_dt, "…", "…", "…", token)
                        naver_results.append((idx, code, lo, hi, div, div_yield, div_dt, close_price))
                    except Exception:
                        pass
        except Exception:
            pass

        if token != self._load_token or not naver_results:
            return

        # Phase 2: Fetch yfinance details in parallel
        def work_yfinance(res):
            idx, code, lo, hi, div, div_yield, div_dt, close_price = res
            div, div_yield, div_dt, div_y1, div_y2, div_y3 = _fetch_yfinance_dividends(code, close_price, div, div_yield, div_dt)
            return idx, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3

        try:
            with ThreadPoolExecutor(max_workers=3) as ex:
                futs = [ex.submit(work_yfinance, r) for r in naver_results]
                for fut in as_completed(futs):
                    if token != self._load_token:
                        return
                    try:
                        idx, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3 = fut.result()
                        # Update the stock dict with final values
                        self._stocks[idx]["_low52"] = lo
                        self._stocks[idx]["_high52"] = hi
                        self._stocks[idx]["_dividend"] = div
                        self._stocks[idx]["_div_yield"] = div_yield
                        self._stocks[idx]["_div_dt"] = div_dt
                        self._stocks[idx]["_div_y1"] = div_y1
                        self._stocks[idx]["_div_y2"] = div_y2
                        self._stocks[idx]["_div_y3"] = div_y3
                        
                        # Cache the result
                        self._details_cache[code] = {
                            "lo": lo, "hi": hi,
                            "div": div, "div_yield": div_yield, "div_dt": div_dt,
                            "div_y1": div_y1, "div_y2": div_y2, "div_y3": div_y3
                        }
                        
                        # Update the UI row with final values
                        self.root.after(0, self._update_details_row, idx, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3, token)
                    except Exception:
                        pass
        except Exception:
            pass

    # ──────────────────────────────────────────────
    # 선택 헬퍼
    # ──────────────────────────────────────────────
    def _select_top5(self):
        children = self.tree.get_children()
        self.tree.selection_set(children[:5])

    def _search_stocks(self):
        """Filter displayed stocks by name keyword.
        If the search box is empty, reset to the full list.
        """
        keyword = self.stock_search_var.get().strip().lower()
        if not keyword:
            self._stocks = list(self._all_stocks)
            if self._current_group:
                self._populate_tree(self._current_group)
            else:
                self._populate_tree(None)
            return
        filtered = [s for s in self._all_stocks if keyword in s.get('stockName', '').lower()]
        self._stocks = filtered
        self._populate_tree(self._current_group)

    def _global_search_stocks_thread(self):
        keyword = self.stock_search_var.get().strip()
        if not keyword:
            messagebox.showwarning("검색 오류", "검색어를 입력해 주세요.")
            return
        
        self.global_search_btn.configure(state="disabled", text="검색 중...")
        self.stats_lbl.configure(text="종목 검색 중...", fg=self.text_muted)
        
        threading.Thread(target=self._global_search_stocks, args=(keyword,), daemon=True).start()

    def _global_search_stocks(self, keyword: str):
        try:
            # 캐시가 빈 경우 동기 구축 시도
            if not getattr(self, "_global_stocks_cache", None):
                self._set_stats("종목 정보 색인 구축 중...", self.text_muted)
                self._build_global_stocks_cache()
                if not getattr(self, "_global_stocks_cache", None):
                    self.root.after(0, lambda: messagebox.showwarning(
                        "검색 대기", "종목 정보를 불러오는 중입니다. 잠시 후 다시 시도해 주세요."))
                    self.root.after(0, lambda: self.stats_lbl.configure(text="대기 필요", fg=self.text_dark))
                    return

            keyword_lower = keyword.lower()
            matched_stocks = []
            
            # 한글 초성 검색 판단
            has_jamo = any(0x3131 <= ord(c) <= 0x314E for c in keyword)
            has_syllable = any(0xAC00 <= ord(c) <= 0xD7A3 for c in keyword)
            chosung_mode = has_jamo and not has_syllable
            qc = _chosung(keyword)
            
            for code, name in self._global_stocks_cache.items():
                name_lower = name.lower()
                matched = False
                if chosung_mode:
                    if qc in _chosung(name):
                        matched = True
                else:
                    if keyword_lower in name_lower or keyword_lower == code:
                        matched = True
                if matched:
                    matched_stocks.append((name, code))
            
            if not matched_stocks:
                self.root.after(0, lambda: messagebox.showinfo("검색 결과", f"'{keyword}'에 대한 검색 결과가 없습니다."))
                self.root.after(0, lambda: self.stats_lbl.configure(text="검색 결과 없음", fg=self.text_dark))
                return
            
            # Limit to top 30 to avoid rate limit/performance issues
            matched_stocks = matched_stocks[:30]
            
            self._set_stats(f"종목 정보 가져오는 중 (0/{len(matched_stocks)})...", self.text_muted)
            
            final_stocks = []
            completed_count = 0
            
            def fetch_one(name, code):
                nonlocal completed_count
                res = _fetch_global_stock_info(code, name)
                completed_count += 1
                self._set_stats(f"종목 정보 가져오는 중 ({completed_count}/{len(matched_stocks)})...", self.text_muted)
                return res
            
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = [executor.submit(fetch_one, name, code) for name, code in matched_stocks]
                for fut in as_completed(futures):
                    try:
                        final_stocks.append(fut.result())
                    except Exception:
                        pass
            
            # Sort by marketValueRaw descending
            final_stocks = sorted(
                final_stocks, 
                key=lambda s: int(s.get("marketValueRaw", 0)) if isinstance(s.get("marketValueRaw", 0), int) or str(s.get("marketValueRaw", 0)).isdigit() else 0, 
                reverse=True
            )
            
            self._stocks = final_stocks
            self._all_stocks = list(final_stocks)
            self._current_group = None
            self._sort_col = None
            self._sort_reverse = False
            
            # Update UI
            self.root.after(0, lambda: self._populate_tree(group=None))
            
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("검색 오류", f"검색 중 오류가 발생했습니다:\n{msg}"))
            self.root.after(0, lambda msg=err_msg: self.stats_lbl.configure(text=f"오류: {msg}", fg="#E63B2E"))
        finally:
            self.root.after(0, lambda: self.global_search_btn.configure(state="normal", text="🔍 전체 검색 (Ctrl+Enter)"))

    def _select_all(self):
        self.tree.selection_set(self.tree.get_children())

    def _deselect_all(self):
        self.tree.selection_remove(self.tree.get_children())

    # ──────────────────────────────────────────────
    # 엑셀(.xlsx) / CSV 저장
    # ──────────────────────────────────────────────
    def _export(self, scope: str):
        if scope == "current":
            if not self._stocks:
                messagebox.showwarning("저장 불가", "저장할 종목이 없습니다.")
                return
            if self._current_group:
                default = f"{self._current_group['name']}_종목.xlsx"
            else:
                default = "검색결과_종목.xlsx"
        else:
            if not self._groups:
                messagebox.showwarning("저장 불가", "업종 목록을 먼저 불러와 주세요.")
                return
            default = "전체업종_종목.xlsx"

        path = filedialog.asksaveasfilename(
            title="엑셀로 저장",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel 통합 문서", "*.xlsx"), ("CSV (UTF-8)", "*.csv")],
        )
        if not path:
            return
        self.export_mb.configure(state="disabled")
        threading.Thread(target=self._export_worker,
             args=(scope, path), daemon=True).start()

    def _export_worker(self, scope: str, path: str):
        try:
            if scope == "current":
                self._attach_details(self._stocks)
                group_name = self._current_group["name"] if self._current_group else "검색결과"
                rows  = self._rows_for_stocks(self._stocks, group_name)
                title = group_name
            else:
                rows = []
                total = len(self._groups)
                for gi, g in enumerate(self._groups):
                    self._set_stats(
                        f"전체 저장 중… 업종 {gi+1}/{total}  ({g['name']})",
                        self.text_muted)
                    try:
                        stocks = _fetch_all_sector_stocks(g["no"])
                    except Exception:
                        continue
                    self._attach_details(stocks)
                    rows += self._rows_for_stocks(stocks, g["name"])
                title = "전체 업종"

            self._write_file(path, rows)
            self.root.after(0, lambda: messagebox.showinfo(
                "저장 완료",
                f"{title} 종목 {len(rows):,}건을 저장했습니다.\n\n{path}"))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(
                "저장 오류", f"엑셀 저장에 실패했습니다:\n{e}"))
        finally:
            self.root.after(0, lambda: self.export_mb.configure(state="normal"))
            self._set_stats_for_current()

    def _attach_details(self, stocks: list):
        # 1. Fill from cache first
        for i, s in enumerate(stocks):
            code = s.get("itemCode", "")
            if code in self._details_cache:
                c = self._details_cache[code]
                s["_low52"] = c["lo"]
                s["_high52"] = c["hi"]
                s["_dividend"] = c["div"]
                s["_div_yield"] = c["div_yield"]
                s["_div_dt"] = c["div_dt"]
                s["_div_y1"] = c["div_y1"]
                s["_div_y2"] = c["div_y2"]
                s["_div_y3"] = c["div_y3"]

        targets = [(i, s) for i, s in enumerate(stocks) if "_low52" not in s]
        if not targets:
            return

        def work(i_s):
            i, s = i_s
            code = s.get("itemCode", "")
            lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3 = _fetch_details(code)
            return i, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3

        with ThreadPoolExecutor(max_workers=12) as ex:
            for i, code, lo, hi, div, div_yield, div_dt, div_y1, div_y2, div_y3 in ex.map(work, targets):
                stocks[i]["_low52"]  = lo
                stocks[i]["_high52"] = hi
                stocks[i]["_dividend"] = div
                stocks[i]["_div_yield"] = div_yield
                stocks[i]["_div_dt"] = div_dt
                stocks[i]["_div_y1"] = div_y1
                stocks[i]["_div_y2"] = div_y2
                stocks[i]["_div_y3"] = div_y3
                
                # Cache it
                self._details_cache[code] = {
                    "lo": lo, "hi": hi,
                    "div": div, "div_yield": div_yield, "div_dt": div_dt,
                    "div_y1": div_y1, "div_y2": div_y2, "div_y3": div_y3
                }

    def _rows_for_stocks(self, stocks: list, sector_name: str) -> list:
        rows = []
        for i, s in enumerate(stocks):
            rows.append([
                i + 1,
                s.get("stockName", ""),
                s.get("itemCode", ""),
                sector_name,
                _fmt_mktcap(int(s.get("marketValueRaw", 0))),
                s.get("closePrice", ""),
                s.get("fluctuationsRatio", ""),
                s.get("_low52", "-"),
                s.get("_high52", "-"),
                s.get("_dividend", "-"),
                s.get("_div_yield", "-"),
                s.get("_div_dt", "-"),
                s.get("_div_y1", "-"),
                s.get("_div_y2", "-"),
                s.get("_div_y3", "-"),
                _ex_code(s),
            ])
        return rows

    def _write_file(self, path: str, rows: list):
        if path.lower().endswith(".csv"):
            self._write_csv(path, rows)
            return
        try:
            self._write_xlsx(path, rows)
        except ImportError:
            alt = path[:-5] + ".csv" if path.lower().endswith(".xlsx") else path + ".csv"
            self._write_csv(alt, rows)

    def _write_csv(self, path: str, rows: list):
        # Excel에서 열었을 때 종목코드(3번째 컬럼)의 leading zero가 유지되도록 ="005930" 형식으로 저장
        formatted_rows = []
        for r in rows:
            new_row = list(r)
            if len(new_row) > 2:
                code = new_row[2]
                new_row[2] = f'="{code}"'
            formatted_rows.append(new_row)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(EXPORT_HEADER)
            w.writerows(formatted_rows)

    def _write_xlsx(self, path: str, rows: list):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "업종별 종목"

        ws.append(EXPORT_HEADER)
        for r in rows:
            ws.append(r)

        # 종목코드 컬럼(C열, index 3)을 텍스트 형식(@)으로 명시적 지정하여 leading zero 보존
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            cell.number_format = '@'
            if cell.value is not None:
                cell.value = str(cell.value)

        head_fill = PatternFill("solid", fgColor="1A1A1A")
        head_font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
        thin      = Side(style="thin", color="D9D9D9")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill      = head_fill
            cell.font      = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        widths = [6, 22, 10, 18, 12, 12, 11, 12, 12, 12, 12, 12, 12, 12, 12, 9]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64 + len(EXPORT_HEADER))}{len(rows) + 1}"
        wb.save(path)

    # ──────────────────────────────────────────────
    # 상태표시 헬퍼(스레드 → UI)
    # ──────────────────────────────────────────────
    def _set_stats(self, text: str, color: str):
        self._safe_after(0, lambda: self.stats_lbl.configure(text=text, fg=color) if hasattr(self, "stats_lbl") and self.stats_lbl and self.stats_lbl.winfo_exists() else None)

    def _set_stats_for_current(self):
        if self._current_group:
            self._safe_after(0, lambda: self._populate_stats(self._current_group))

    def _populate_stats(self, group: dict):
        try:
            if not hasattr(self, "stats_lbl") or not self.stats_lbl or not self.stats_lbl.winfo_exists():
                return
            cr    = float(group.get("changeRate", 0))
            sign  = "▲" if cr >= 0 else "▼"
            color = "#E63B2E" if cr >= 0 else "#0055FF"
            self.stats_lbl.configure(
                fg=color,
                text=(
                    f"{sign} {abs(cr):.2f}%  |  "
                    f"상승 {group['riseCount']}  하락 {group['fallCount']}  "
                    f"보합 {group['steadyCount']}  (총 {group['totalCount']}종목)"
                ),
            )
        except Exception:
            pass

    # ──────────────────────────────────────────────
    # 선택 종목 차트 보기 연동
    # ──────────────────────────────────────────────
    def _show_chart(self):
        selected = self.tree.selection()
        if not selected:
            # Default to top 5 stocks
            children = self.tree.get_children()
            if not children:
                messagebox.showwarning("조회 불가", "조회할 종목이 없습니다. 먼저 업종을 선택해 주세요.")
                return
            selected = children[:5]

        stocks_to_show = []
        for iid in selected:
            s = self._stocks[int(iid)]
            code = s.get("itemCode", "")
            name = s.get("stockName", "")
            ex = _ex_code(s)
            symbol = f"{code}.{ex}"
            stocks_to_show.append((symbol, name))
            
        if not stocks_to_show:
            return

        from chart_viewer import ChartViewer
        ChartViewer(self.root, stocks_to_show, self.theme)

    # ──────────────────────────────────────────────
    # 배당수익률 정렬 헬퍼
    # ──────────────────────────────────────────────
    def _toggle_dividend_yield_sort(self):
        if self._sort_col == "배당수익률":
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = "배당수익률"
            self._sort_reverse = True  # 첫 클릭 시 내림차순(고배당순)

        self._stocks = sorted(
            self._stocks,
            key=lambda s: self._get_dividend_yield_sort_key(s, self._sort_reverse),
            reverse=self._sort_reverse
        )

        if self._current_group:
            self._populate_tree(self._current_group)

    def _get_dividend_yield_sort_key(self, s, reverse):
        val = s.get("_div_yield", "-")
        if not val or val in ("-", "…", "N/A"):
            return float('-inf') if reverse else float('inf')
        try:
            return float(val.replace("%", "").strip())
        except ValueError:
            return float('-inf') if reverse else float('inf')

    def _update_headers(self):
        col_names = {
            "순위": "순위",
            "종목명": "종목명",
            "종목코드": "종목코드",
            "시가총액": "시가총액",
            "종가": "종가",
            "등락률": "등락률",
            "52주최저": "52주최저",
            "52주최고": "52주최고",
            "배당금": "배당금",
            "배당수익률": "배당수익률",
            "배당기준일": "배당기준일",
            "1년전 배당": "1년전 배당",
            "2년전 배당": "2년전 배당",
            "3년전 배당": "3년전 배당",
            "거래소": "거래소"
        }
        for col_id, base_text in col_names.items():
            if col_id == "배당수익률":
                if col_id == self._sort_col:
                    arrow = " ▼" if self._sort_reverse else " ▲"
                    self.tree.heading(col_id, text=base_text + arrow, command=self._toggle_dividend_yield_sort)
                else:
                    self.tree.heading(col_id, text=base_text, command=self._toggle_dividend_yield_sort)
            else:
                if col_id == self._sort_col:
                    arrow = " ▼" if self._sort_reverse else " ▲"
                    self.tree.heading(col_id, text=base_text + arrow)
                else:
                    self.tree.heading(col_id, text=base_text)

